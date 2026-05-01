"""
simulation_service.py
──────────────────────
MOC transient simulation + thermal-hydraulic model for Alaska pipeline.

Thermal model (operator-splitting):
  Every broadcast step:
  1. MOC computes pressure/flow (30 time steps)
  2. Thermal model computes temperature profile along all 7 pipe segments
     using pipe_temperature_loss() + apply_crude_heater() from
     crude_thermal_hydraulic.py
  3. Updated viscosity → new friction factor per segment → pushed into MOC

Broadcast payload includes:
  temperature_profile  : [{km, temp_c, node}] — for distance-temperature chart
  heater_states        : {HS1: {on, target_c, actual_mw, ...}, HS2: {...}}
  node_temperatures    : {PS1: temp_c, PS3: temp_c, ..., Sink: temp_c}
"""

from __future__ import annotations
import asyncio, os, threading, time, copy, math
from datetime import datetime
from typing import Dict, List, Optional

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from moc_solver import MOCSolver
from networks.alaska_pipeline import build_alaska_network, DEFAULT_HEATERS, THERMAL_DEFAULTS
from crude_oil_temperature_library import CrudeOilTemperatureLibrary, CrudeOilTemperatureConfig
from dra_engine import DRAEngine
from leak_detection import LeakDetector
from crude_thermal_hydraulic import (
    PipeSection as ThPipeSection,
    HeaterStation as ThHeaterStation,
    pipe_temperature_loss,
    apply_crude_heater,
    crude_viscosity_from_temperature,
)

import numpy as np

def sanitize(data):
    if isinstance(data, dict):
        return {k: sanitize(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize(v) for v in data]
    elif isinstance(data, np.bool_):
        return bool(data)
    elif isinstance(data, np.integer):
        return int(data)
    elif isinstance(data, np.floating):
        return float(data)
    else:
        return data

MOC_STEPS_PER_BROADCAST = 30
PRESSURE_NOISE_STD       = 0.005
FLOW_NOISE_STD           = 0.002
EXCEL_FLUSH_EVERY        = 50

# Segment lengths (must match alaska_pipeline.py exactly)
_SEG_LENGTHS = {
    "PS1_PS3" : 132_000,
    "PS3_PS4" :  64_000,
    "PS4_PS5" :  78_000,
    "PS5_HS1" : 200_000,
    "HS1_PS9" : 202_000,
    "PS9_HS2" : 300_000,
    "HS2_Sink": 312_000,
}

# Which heater (if any) sits at the end of each segment
_SEG_HEATER = {
    "PS5_HS1": "HS1",
    "PS9_HS2": "HS2",
}

# Node name at end of each segment (for profile labelling)
_SEG_END_NODE = {
    "PS1_PS3" : "PS3",
    "PS3_PS4" : "PS4",
    "PS4_PS5" : "PS5",
    "PS5_HS1" : "HS1_in",
    "HS1_PS9" : "PS9",
    "PS9_HS2" : "HS2_in",
    "HS2_Sink": "Sink",
}


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _make_header(node_names, pipe_names, pump_names):
    c  = ["step", "simulated_time_s"]
    c += [f"{n}_bar"        for n in node_names]
    c += [f"{n}_sensor_bar" for n in node_names]
    c += [f"{n}_m3s"        for n in pipe_names]
    c += [f"{n}_tempC"      for n in pipe_names]
    c += [f"{n}_friction_f" for n in pipe_names]
    c += [f"{n}_on"         for n in pump_names]
    c += [f"{n}_alpha"      for n in pump_names]
    c += ["HS1_on","HS1_target_c","HS1_actual_mw"]
    c += ["HS2_on","HS2_target_c","HS2_actual_mw"]
    c += ["ambient_temp_c","inlet_temp_c","leak_nodes"]
    return c


def _make_row(step, sim_t, calc_p, sensor_p, flows, pipe_temps, pipe_f,
              pump_st, node_names, pipe_names, pump_names,
              hs, ambient_c, inlet_c, leaks):
    r  = [step, round(sim_t, 2)]
    r += [calc_p.get(n,0)   for n in node_names]
    r += [sensor_p.get(n,0) for n in node_names]
    r += [flows.get(n,0)    for n in pipe_names]
    r += [round(pipe_temps.get(n,29),2) for n in pipe_names]
    r += [round(pipe_f.get(n,0.01),6)   for n in pipe_names]
    r += [int(pump_st.get(n,{}).get("on",False))   for n in pump_names]
    r += [round(pump_st.get(n,{}).get("alpha",0),4) for n in pump_names]
    r += [int(hs.get("HS1",{}).get("on",True)), hs.get("HS1",{}).get("target_c",40), hs.get("HS1",{}).get("actual_mw",0)]
    r += [int(hs.get("HS2",{}).get("on",True)), hs.get("HS2",{}).get("target_c",35), hs.get("HS2",{}).get("actual_mw",0)]
    r += [ambient_c, inlet_c, "|".join(leaks.keys()) if leaks else ""]
    return r


def _create_workbook(header):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Simulation"
    hfill = PatternFill("solid", fgColor="EFF2F7")
    hfont = Font(name="Consolas", bold=True, color="2B7FC4", size=9)
    for ci, col in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(col)+2, 12)
    ws.freeze_panes = "A2"
    return wb


# ── Simulation service ────────────────────────────────────────────────────────

class SimulationService:

    def __init__(self, sio=None) -> None:
        self.sio     = sio
        self.running = False
        self._thread: Optional[threading.Thread] = None
        self._lock   = threading.Lock()

        self.solver: Optional[MOCSolver] = None
        self.info:   Optional[dict]      = None

        self._leaks:             Dict[str, float] = {}  # named node leaks: node -> kg/s
        self._manual_leaks:      Dict[str, dict] = {}   # arbitrary km leaks: leak_id -> metadata
        self._paused:            bool = False
        self._safety_interlock:  bool = True   # auto-trip upstream pump on valve close
        self._history_maxlen = 500
        self.history: list   = []
        self.step_counter    = 0

        # Crude oil library for Colebrook friction
        self._crude_lib = CrudeOilTemperatureLibrary(
            CrudeOilTemperatureConfig(api_gravity=33.4, pipe_diameter_m=1.2192,
                                      pipe_roughness_m=4.6e-5, default_mass_flow_kg_s=800.0)
        )

        # Thermal state
        td = THERMAL_DEFAULTS
        self._inlet_temp_c:   float = td["inlet_temp_c"]
        self._ambient_temp_c: float = td["ambient_temp_c"]
        self._U:              float = td["U_W_m2K"]
        self._cp:             float = td["cp_J_kgK"]
        self._mu_ref:         float = td["mu_ref_Pa_s"]
        self._T_ref_K:        float = td["T_ref_K"]
        self._B_K:            float = td["B_K"]

        # Heater configs (deep copy so network defaults are not mutated)
        self._heater_cfg: Dict[str, dict] = copy.deepcopy(DEFAULT_HEATERS)
        # Live heater output state
        self._heater_states: Dict[str, dict] = {
            k: {"on": v["on"], "target_c": v["target_c"], "max_mw": v["max_mw"],
                "efficiency": v["efficiency"], "actual_mw": 0.0, "fuel_mw": 0.0, "active": False}
            for k, v in self._heater_cfg.items()
        }

        # Per-segment temps and friction
        self._pipe_temps:             Dict[str, float] = {}
        self._pipe_friction:          Dict[str, float] = {}   # f_effective (with DRA)
        self._pipe_friction_thermal:  Dict[str, float] = {}   # f_thermal (no DRA)

        # Valve states and closure schedules
        # Valves sit BEFORE PS3, PS4, PS9 (not before PS1)
        self._valve_names    = ["V_PS3", "V_PS4", "V_PS9"]
        self._valve_junction = {"V_PS3": "PS3", "V_PS4": "PS4", "V_PS9": "PS9"}
        # Each valve: {tau, state, schedule}
        # state: 'open'|'closing'|'opening'|'closed'
        # schedule: {start_t, end_t, tau_start, tau_end} or None
        self._valve_states: dict = {
            v: {"tau": 1.0, "state": "open", "schedule": None}
            for v in self._valve_names
        }

        # DRA engine — manages per-pump injection and friction reduction
        self._dra = DRAEngine()

        # Leak detection layer: EWMA + CUSUM + negative-pressure-wave + flow balance
        self._leak_detector = LeakDetector()

        # Latest temperature profile for broadcast
        self._temp_profile: List[dict] = []
        self._node_temps:   Dict[str, float] = {}

        # Excel
        self._wb, self._ws, self._excel_path = None, None, ""
        self._excel_header: List[str] = []
        self._row_buffer:   List[List] = []

    # ── Lifecycle ──────────────────────────────────────────────────────────────

    def start(self) -> None:
        if self.running: print("⚠️ Already running"); return

        print("\n" + "="*56)
        print("   ALASKA PIPELINE — MOC + THERMAL-HYDRAULIC")
        print("="*56)

        moc_config, self.info = build_alaska_network()
        self.solver = MOCSolver(moc_config)
        self.solver.start_all_pumps()

        # Initial thermal profile
        self._update_thermal_profile(mass_flow_kg_s=800.0)

        print(f"\n   Δt = {self.solver.dt:.4f} s  |  {len(self.solver.pipes)} pipe segments")
        print(f"   Inlet temp   : {self._inlet_temp_c}°C")
        print(f"   Ambient temp : {self._ambient_temp_c}°C")
        print(f"   Heaters      : HS1={self._heater_cfg['HS1']['target_c']}°C/{self._heater_cfg['HS1']['max_mw']}MW  "
              f"HS2={self._heater_cfg['HS2']['target_c']}°C/{self._heater_cfg['HS2']['max_mw']}MW\n")

        self._init_excel()
        self.running = True; self.step_counter = 0
        self.history.clear(); self._leaks.clear(); self._manual_leaks.clear(); self._row_buffer.clear()
        self._leak_detector.reset()

        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        print(f"   ✅ Simulation started  |  Excel: {self._excel_path}\n")

    def stop(self) -> None:
        if not self.running: return
        self.running = False
        if self._thread: self._thread.join(timeout=3)
        self._flush_excel(final=True)
        print(f"\n🛑 Stopped  steps={self.step_counter}")

    # ── Thermal model ─────────────────────────────────────────────────────────

    def _update_thermal_profile(self, mass_flow_kg_s: float) -> None:
        """
        Compute temperature at every node along the pipeline.
        Uses pipe_temperature_loss (exponential cooling) and apply_crude_heater.
        Updates self._pipe_temps, self._pipe_friction, self._temp_profile, self._node_temps.
        """
        T_K   = self._inlet_temp_c + 273.15
        T_amb = self._ambient_temp_c + 273.15
        D     = self.info["diameter_m"]
        rough = THERMAL_DEFAULTS["roughness_m"]
        km    = 0.0

        profile = [{"km": 0.0, "temp_c": self._inlet_temp_c, "node": "PS1"}]
        node_temps = {"PS1": self._inlet_temp_c}

        for seg_name, seg_len in _SEG_LENGTHS.items():
            # Pipe heat loss
            th_pipe = ThPipeSection(seg_name, seg_len, D, rough, self._U)
            T_out_K = pipe_temperature_loss(T_K, T_amb, mass_flow_kg_s, self._cp, th_pipe)

            km += seg_len / 1000.0
            end_node = _SEG_END_NODE[seg_name]
            T_out_C  = round(T_out_K - 273.15, 2)
            node_temps[end_node] = T_out_C
            profile.append({"km": round(km, 1), "temp_c": T_out_C, "node": end_node})

            # Segment mid-point temperature for property averaging
            T_avg_K = (T_K + T_out_K) / 2.0
            snap = self._crude_lib.property_snapshot(temp_c=T_avg_K - 273.15,
                                                     mass_flow_kg_s=mass_flow_kg_s)
            f_new = snap["darcy_friction_factor"] or THERMAL_DEFAULTS.get("f_default", 0.01)
            rho_new = snap["density_kg_m3"]

            self._pipe_temps[seg_name]             = T_out_C
            self._pipe_friction_thermal[seg_name] = f_new   # thermal only, before DRA

            # Apply DRA friction reduction on top of thermal friction
            if self._dra.any_active:
                dra_map = self._dra.compute_effective_friction({seg_name: f_new})
                f_effective = dra_map[seg_name][0]
            else:
                f_effective = f_new
            self._pipe_friction[seg_name] = f_effective

            # Push into MOC solver
            if self.solver:
                self.solver.update_pipe_temperature(seg_name, f_effective, rho_new, T_out_C)

            # Apply heater at end of segment if present
            heater_id = _SEG_HEATER.get(seg_name)
            if heater_id:
                hs = self._heater_cfg[heater_id]
                if hs["on"]:
                    heater = ThHeaterStation(
                        id=heater_id,
                        target_temperature_K=hs["target_c"] + 273.15,
                        max_heat_power_W=hs["max_mw"] * 1e6,
                        efficiency=hs["efficiency"],
                    )
                    h_result = apply_crude_heater(T_out_K, mass_flow_kg_s, self._cp, heater)
                    T_out_K = h_result["T_out_K"]
                    self._heater_states[heater_id].update({
                        "actual_mw": round(h_result["Q_useful_W"] / 1e6, 2),
                        "fuel_mw"  : round(h_result["Q_fuel_W"]    / 1e6, 2),
                        "active"   : h_result["heater_active"],
                    })
                    out_c = round(T_out_K - 273.15, 2)
                    out_node = heater_id + "_out"
                    node_temps[out_node] = out_c
                    profile.append({"km": round(km, 1), "temp_c": out_c, "node": out_node})
                else:
                    self._heater_states[heater_id].update({"actual_mw":0, "fuel_mw":0, "active":False})

            T_K = T_out_K

        # Map HS1_out / HS2_out → HS1 / HS2 keys for the frontend
        for hs_id in ("HS1", "HS2"):
            if hs_id + "_out" in node_temps:
                node_temps[hs_id] = node_temps[hs_id + "_out"]
            elif hs_id + "_in" in node_temps:
                node_temps[hs_id] = node_temps[hs_id + "_in"]

        self._temp_profile = profile
        self._node_temps   = node_temps

    # ── Heater control ────────────────────────────────────────────────────────

    def set_heater(self, heater_id: str, target_c: float = None, max_mw: float = None, on: bool = None) -> dict:
        if heater_id not in self._heater_cfg:
            return {"error": f"Unknown heater '{heater_id}'. Available: {list(self._heater_cfg.keys())}"}
        with self._lock:
            if target_c is not None:
                self._heater_cfg[heater_id]["target_c"] = float(target_c)
                self._heater_states[heater_id]["target_c"] = float(target_c)
            if max_mw is not None:
                self._heater_cfg[heater_id]["max_mw"] = float(max_mw)
                self._heater_states[heater_id]["max_mw"] = float(max_mw)
            if on is not None:
                self._heater_cfg[heater_id]["on"] = bool(on)
                self._heater_states[heater_id]["on"] = bool(on)
        print(f"🔥 {heater_id}: target={self._heater_cfg[heater_id]['target_c']}°C  "
              f"max={self._heater_cfg[heater_id]['max_mw']}MW  "
              f"on={self._heater_cfg[heater_id]['on']}")
        return {"status": "ok", "heater": heater_id, "config": dict(self._heater_cfg[heater_id])}

    def get_heater_states(self) -> Dict[str, dict]:
        return {k: {**self._heater_cfg[k], **self._heater_states[k]} for k in self._heater_cfg}

    def set_ambient_temp(self, temp_c: float) -> str:
        with self._lock: self._ambient_temp_c = float(temp_c)
        return f"✅ Ambient temperature set to {temp_c}°C"

    def set_inlet_temp(self, temp_c: float) -> str:
        with self._lock: self._inlet_temp_c = float(temp_c)
        return f"✅ Inlet temperature set to {temp_c}°C"

    # ── Pump control ──────────────────────────────────────────────────────────

    def pump_start(self, name: str) -> str:
        if not self.info or name not in self.info["pump_stations"]: return f"❌ Unknown pump '{name}'"
        with self._lock:
            if self.solver: self.solver.start_pump(name)
        print(f"🟢 PUMP START: {name}"); return f"✅ {name} ramping up"

    def pump_stop(self, name: str) -> str:
        if not self.info or name not in self.info["pump_stations"]: return f"❌ Unknown pump '{name}'"
        with self._lock:
            if self.solver: self.solver.stop_pump(name)
        print(f"🔴 PUMP TRIP: {name}"); return f"✅ {name} tripped"


    # ── Pause / resume ───────────────────────────────────────────────────────────

    def pause(self) -> str:
        self._paused = True
        print("⏸  Simulation PAUSED")
        return "✅ Simulation paused"

    def resume(self) -> str:
        self._paused = False
        print("▶  Simulation RESUMED")
        return "✅ Simulation resumed"

    def set_interlock(self, enabled: bool) -> str:
        self._safety_interlock = bool(enabled)
        state = "ENABLED" if enabled else "DISABLED"
        print(f"🔒 Safety interlock {state}")
        return f"✅ Safety interlock {state}"

    def get_interlock(self) -> bool:
        return self._safety_interlock

    # ── Valve control ────────────────────────────────────────────────────────

    VALVE_CLOSE_TIME = 60.0   # seconds for full closure (realistic large pipeline valve)
    VALVE_OPEN_TIME  = 90.0   # seconds for full opening

    # Upstream pump for each valve (nearest pump feeding into the valve)
    _VALVE_UPSTREAM_PUMP = {
        "V_PS3": "PS1",   # PS1 feeds PS1_PS3 → V_PS3
        "V_PS4": "PS3",   # PS3 feeds PS3_PS4 → V_PS4
        "V_PS9": "PS4",   # PS4 feeds through PS5/HS1 → V_PS9 (no pump between PS4 and PS9)
    }

    def close_valve(self, valve_name: str, duration_s: float = None) -> str:
        """
        Start closing a valve over duration_s seconds.
        The MOC solver sees tau decrease each step → water hammer propagates naturally.
        Duration defaults to VALVE_CLOSE_TIME (60 s).
        """
        if valve_name not in self._valve_states:
            return f"❌ Unknown valve '{valve_name}'. Available: {self._valve_names}"
        dur = float(duration_s) if duration_s else self.VALVE_CLOSE_TIME
        sim_t = self.solver.simulation_time if self.solver else 0.0
        with self._lock:
            tau_now = self._valve_states[valve_name]["tau"]
            self._valve_states[valve_name].update({
                "state"   : "closing",
                "schedule": {
                    "start_t"  : sim_t,
                    "end_t"    : sim_t + dur,
                    "tau_start": tau_now,
                    "tau_end"  : 0.0,
                },
            })
        interlock_msg = ""
        if self._safety_interlock:
            upstream = self._VALVE_UPSTREAM_PUMP.get(valve_name)
            if upstream:
                self.pump_stop(upstream)
                interlock_msg = f" | Interlock: {upstream} tripped"
        print(f"🔴 VALVE {valve_name}: closing over {dur:.0f}s from t={sim_t:.1f}s{interlock_msg}")
        return f"✅ {valve_name} closing over {dur:.0f} s — water hammer will propagate{interlock_msg}"

    def open_valve(self, valve_name: str, duration_s: float = None) -> str:
        """Start opening a valve over duration_s seconds."""
        if valve_name not in self._valve_states:
            return f"❌ Unknown valve '{valve_name}'. Available: {self._valve_names}"
        dur = float(duration_s) if duration_s else self.VALVE_OPEN_TIME
        sim_t = self.solver.simulation_time if self.solver else 0.0
        with self._lock:
            tau_now = self._valve_states[valve_name]["tau"]
            self._valve_states[valve_name].update({
                "state"   : "opening",
                "schedule": {
                    "start_t"  : sim_t,
                    "end_t"    : sim_t + dur,
                    "tau_start": tau_now,
                    "tau_end"  : 1.0,
                },
            })
        print(f"🟢 VALVE {valve_name}: opening over {dur:.0f}s from t={sim_t:.1f}s")
        return f"✅ {valve_name} opening over {dur:.0f} s"

    def get_valve_states(self) -> dict:
        return {
            v: {
                "tau"     : round(s["tau"], 4),
                "state"   : s["state"],
                "junction": self._valve_junction[v],
                "pct_open": round(s["tau"] * 100, 1),
            }
            for v, s in self._valve_states.items()
        }

    def _update_valves(self, sim_t: float) -> None:
        """
        Called every broadcast step to advance valve schedules.
        Pushes updated tau into the MOC solver.
        """
        for v_name, vs in self._valve_states.items():
            sched = vs.get("schedule")
            if sched is None:
                continue
            t0, t1 = sched["start_t"], sched["end_t"]
            if sim_t >= t1:
                # Schedule complete
                vs["tau"]      = sched["tau_end"]
                vs["state"]    = "closed" if sched["tau_end"] < 0.001 else "open"
                vs["schedule"] = None
            elif sim_t >= t0:
                frac      = (sim_t - t0) / (t1 - t0)
                vs["tau"] = sched["tau_start"] + frac * (sched["tau_end"] - sched["tau_start"])
                # state stays 'closing' or 'opening'

            # Push into MOC solver
            if self.solver:
                junction = self._valve_junction[v_name]
                self.solver.set_valve_tau(junction, vs["tau"])

    # ── DRA control ──────────────────────────────────────────────────────────

    def set_dra(self, station_id: str, on: bool = None, conc_ppm: float = None) -> dict:
        """
        Set DRA injection at a pump station.
        Changes take effect from the next broadcast step.
        station_id : 'PS1', 'PS3', 'PS4', or 'PS9'
        on         : True = injecting, False = stopped
        conc_ppm   : injection concentration 0–250 ppm
        """
        with self._lock:
            result = self._dra.set_injection(station_id, on=on, conc_ppm=conc_ppm)
        if "error" in result:
            return result
        active = [s for s, d in result.items() if d["on"]]
        print(f"💧 DRA: {station_id} on={on} conc={conc_ppm}ppm  active={active}")
        return {"status": "ok", "dra_states": result}

    def clear_all_dra(self) -> str:
        with self._lock:
            self._dra.clear_all()
        print("💧 DRA: all injections cleared")
        return "✅ All DRA injections cleared"

    def get_dra_states(self) -> dict:
        return {
            "injectors"  : self._dra.get_state(),
            "any_active" : self._dra.any_active,
            "segment_profile": self._dra.segment_dra_profile(),
        }

    def get_pump_states(self) -> Dict[str, dict]:
        if not self.solver or not self.info: return {}
        return {n: {"on": self.solver._pump_on.get(n,False),
                    "alpha": round(self.solver._alpha(n),4),
                    "ramp_pct": round(self.solver._alpha(n)*100,1),
                    "at_speed": bool(self.solver._alpha(n) >= 0.999)}
                for n in self.info["pump_stations"]}

    # ── Leak control ──────────────────────────────────────────────────────────

    def _km_to_pipe_cell(self, location_km: float) -> dict:
        """Map a user-entered mainline km to the nearest MOC grid cell."""
        if not self.info:
            raise ValueError("Simulation network is not initialized")
        total_km = float(self.info["node_km"].get("Sink", 1288.0))
        x_km = max(0.0, min(float(location_km), total_km))
        cursor_km = 0.0
        for pipe_idx, seg in enumerate(self.info["pipe_segments"]):
            length_km = float(seg["length_km"])
            start_km = cursor_km
            end_km = cursor_km + length_km
            if start_km <= x_km <= end_km or pipe_idx == len(self.info["pipe_segments"]) - 1:
                local_km = max(0.0, min(length_km, x_km - start_km))
                dx_km = float(self.info.get("dx", 2000.0)) / 1000.0
                node_idx = int(round(local_km / dx_km))
                # Interior cell leaks are clamped away from boundaries in MOCSolver.
                pipe = self.solver.pipes[pipe_idx] if self.solver else None
                if pipe is not None:
                    node_idx = max(1, min(pipe.N - 1, node_idx))
                    snapped_local_km = node_idx * dx_km
                    snapped_km = start_km + snapped_local_km
                else:
                    snapped_km = x_km
                return {
                    "pipe_idx": pipe_idx,
                    "pipe_name": seg["name"],
                    "from": seg["from"],
                    "to": seg["to"],
                    "node_idx": node_idx,
                    "requested_km": round(x_km, 3),
                    "snapped_km": round(snapped_km, 3),
                    "local_km": round(local_km, 3),
                }
            cursor_km = end_km
        raise ValueError("Could not map leak location to pipe segment")

    def _apply_manual_leak_impact(self, sim_t: float, calc_p: Dict[str, float], flows_raw: Dict[str, float]) -> None:
        """Apply a visible, physics-motivated manual-leak effect to UI signals.

        The MOC cell leak is still active in the solver. This overlay fixes the
        previous demo problem where an interior leak could be too local/subtle
        in one-broadcast UI charts, so pressure/flow looked unchanged even when
        the leak object existed.

        Model used here:
          • leak acts as an additional outflow q_leak
          • a negative pressure wave reaches each node after distance/a
          • after arrival, nodes see a sustained pressure loss attenuated by distance
          • downstream segment flows reduce by q_leak; upstream flows rise slightly
        """
        if not self.info or not self._manual_leaks:
            return

        node_km = self.info.get("node_km", {})
        wave_kmps = max(float(self.info.get("wave_speed", 1200.0)) / 1000.0, 1e-9)
        density = max(float(self.info.get("density", 860.0)), 1e-9)
        abs_flows = [abs(float(v)) for v in (flows_raw or {}).values() if v is not None]
        base_flow_m3s = max(sum(abs_flows) / max(len(abs_flows), 1), 0.05)

        # Cumulative pipe start/end km lookup
        seg_bounds = []
        cursor = 0.0
        for seg in self.info.get("pipe_segments", []):
            length = float(seg.get("length_km", 0.0))
            seg_bounds.append((seg.get("name"), cursor, cursor + length))
            cursor += length

        for leak in list(self._manual_leaks.values()):
            try:
                x_leak = float(leak.get("location_km"))
                q_m3s = float(leak.get("flow_m3s", 0.0))
                t0 = float(leak.get("created_at_s", sim_t))
            except (TypeError, ValueError):
                continue
            age = max(0.0, float(sim_t) - t0)
            if q_m3s <= 0.0:
                continue

            # Leak severity relative to normal line flow. Clamp to avoid insane UI values.
            severity = max(0.005, min(q_m3s / base_flow_m3s, 0.35))
            steady_drop_bar = max(0.8, min(9.0, 55.0 * severity))
            wave_pulse_bar = 0.9 * steady_drop_bar

            # Pressure response at key nodes.
            for node, km in node_km.items():
                if node not in calc_p:
                    continue
                dist_km = abs(float(km) - x_leak)
                arrival_delay = dist_km / wave_kmps
                if age < arrival_delay:
                    continue
                since_arrival = age - arrival_delay
                attenuation = math.exp(-dist_km / 520.0)
                # Fast negative wave pulse + slower persistent pressure loss.
                pulse = math.exp(-((since_arrival) / 55.0) ** 2)
                ramp = min(1.0, since_arrival / 180.0)
                drop = (wave_pulse_bar * pulse + steady_drop_bar * ramp) * attenuation
                calc_p[node] = round(max(0.0, float(calc_p[node]) - drop), 4)

            # Flow response. A leak inside a segment should create a clear split:
            # upstream has to supply extra demand; downstream receives less flow.
            for name, start_km, end_km in seg_bounds:
                if not name or name not in flows_raw:
                    continue
                mid_km = 0.5 * (start_km + end_km)
                arrival_delay = abs(mid_km - x_leak) / wave_kmps
                if age < arrival_delay:
                    continue
                ramp = min(1.0, (age - arrival_delay) / 120.0)
                if end_km <= x_leak:
                    # Upstream side sees a small increase because the leak is an added demand.
                    flows_raw[name] = float(flows_raw[name]) + 0.25 * q_m3s * ramp
                elif start_km >= x_leak:
                    # Downstream of the leak, delivered flow reduces by approximately leak outflow.
                    flows_raw[name] = max(0.0, float(flows_raw[name]) - q_m3s * ramp)
                else:
                    # Segment containing the leak: chart plots only one value, so show the average
                    # of upstream/downstream sides instead of hiding the interior loss.
                    flows_raw[name] = max(0.0, float(flows_raw[name]) - 0.55 * q_m3s * ramp)

    def add_leak(self, node: str, flow_kg_s: float) -> str:
        if not self.info or node not in self.info["junction_names"]: return f"❌ Unknown node '{node}'"
        with self._lock:
            self._leaks[node] = float(flow_kg_s)
            if self.solver: self.solver.add_leak(node, float(flow_kg_s)/self.info["density"])
        return f"✅ Leak at {node}: {flow_kg_s:.2f} kg/s"

    def add_leak_at_km(self, location_km: float, flow_kg_s: float) -> dict:
        """Manually inject a leak at any mainline km and snap it to the MOC grid."""
        if not self.info:
            return {"status": "error", "message": "Simulation network is not initialized"}
        if float(flow_kg_s) <= 0:
            return {"status": "error", "message": "Leak flow must be greater than zero"}
        try:
            loc = self._km_to_pipe_cell(location_km)
        except Exception as e:
            return {"status": "error", "message": str(e)}

        leak_id = f"MANUAL_{len(self._manual_leaks)+1:02d}_{int(round(loc['snapped_km']*1000))}m"
        q_m3s = float(flow_kg_s) / float(self.info["density"])
        meta = {
            "id": leak_id,
            "type": "manual_location",
            "location_km": loc["snapped_km"],
            "requested_location_km": loc["requested_km"],
            "pipe_name": loc["pipe_name"],
            "from": loc["from"],
            "to": loc["to"],
            "pipe_idx": loc["pipe_idx"],
            "node_idx": loc["node_idx"],
            "flow_kg_s": round(float(flow_kg_s), 3),
            "flow_m3s": round(q_m3s, 6),
            "created_at_s": round(self.solver.simulation_time, 3) if self.solver else 0.0,
        }
        with self._lock:
            self._manual_leaks[leak_id] = meta
            if self.solver:
                self.solver.add_cell_leak(leak_id, loc["pipe_idx"], loc["node_idx"], q_m3s)
            # Clear old wave evidence but keep healthy baseline/previous sample so the new leak can be detected.
            self._leak_detector.start_new_wave_event()
        print(f"💥 Manual leak: {leak_id} at {meta['location_km']} km on {meta['pipe_name']}  flow={flow_kg_s:.2f} kg/s")
        return {"status": "ok", "message": f"✅ Manual leak injected at {meta['location_km']:.1f} km ({meta['pipe_name']}), {flow_kg_s:.1f} kg/s", "leak": meta}

    def remove_leak(self, node: str) -> str:
        with self._lock:
            if node not in self._leaks: return f"ℹ️ No leak at '{node}'"
            del self._leaks[node]
            if self.solver: self.solver.remove_leak(node)
            self._leak_detector.reset()
        return f"✅ Leak removed at {node}"

    def remove_leak_by_id(self, leak_id: str) -> str:
        with self._lock:
            removed = False
            if leak_id in self._manual_leaks:
                del self._manual_leaks[leak_id]
                if self.solver: self.solver.remove_cell_leak(leak_id)
                removed = True
            if leak_id in self._leaks:
                del self._leaks[leak_id]
                if self.solver: self.solver.remove_leak(leak_id)
                removed = True
            if removed:
                self._leak_detector.reset()
                return f"✅ Leak removed: {leak_id}"
        return f"ℹ️ No leak with id '{leak_id}'"

    def clear_leaks(self) -> str:
        with self._lock:
            self._leaks.clear()
            self._manual_leaks.clear()
            if self.solver: self.solver.clear_leaks()
            self._leak_detector.reset()
        return "✅ All leaks removed"

    def get_leaks(self):
        merged = {node: {"id": node, "type": "node", "node": node, "flow_kg_s": flow} for node, flow in self._leaks.items()}
        merged.update({k: dict(v) for k, v in self._manual_leaks.items()})
        return merged

    def get_status(self) -> dict:
        raw = {"running": self.running, "step": self.step_counter,
               "simulated_time_s": round(float(self.solver.simulation_time), 2) if self.solver else 0.0,
               "active_leaks": self.get_leaks(), "pump_states": self.get_pump_states(),
               "heater_states": self.get_heater_states(),
               "ambient_temp_c": float(self._ambient_temp_c), "inlet_temp_c": float(self._inlet_temp_c),
               "excel_file": self._excel_path}
        return sanitize(raw)

    def get_latest(self): return self.history[-1] if self.history else None

    def get_leak_detection_status(self): return self._leak_detector.get_status()

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _init_excel(self) -> None:
        self._excel_header = _make_header(self.info["junction_names"], self.info["pipe_names"], self.info["pump_stations"])
        self._wb = _create_workbook(self._excel_header); self._ws = self._wb.active
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"simulation_data_{ts}.xlsx")

    def _flush_excel(self, final=False) -> None:
        if not self._wb or not self._row_buffer: return
        try:
            for row in self._row_buffer: self._ws.append(row)
            self._row_buffer.clear(); self._wb.save(self._excel_path)
            if final: print(f"📊 Excel: {self._excel_path} ({self._ws.max_row-1} rows)")
        except Exception as e: print(f"⚠️ Excel error: {e}")

    # ── Simulation loop ────────────────────────────────────────────────────────


    def _build_friction_profile(self) -> list:
        """
        Build friction factor profile list for the distance chart.
        Returns [{km, f_thermal, f_effective}] with 3 points per segment
        (start, mid, end) so the chart covers km=0 to km=1288 correctly.
        Each segment has constant f (based on midpoint temperature and DRA).
        The step-changes at segment boundaries show DRA injection effects.
        """
        SEG_KM = {
            "PS1_PS3" : (0,   132),
            "PS3_PS4" : (132, 196),
            "PS4_PS5" : (196, 274),
            "PS5_HS1" : (274, 474),
            "HS1_PS9" : (474, 676),
            "PS9_HS2" : (676, 976),
            "HS2_Sink": (976, 1288),
        }
        pts = []
        last_f_th, last_f_eff = None, None
        for seg, (km_s, km_e) in SEG_KM.items():
            f_th  = self._pipe_friction_thermal.get(seg, 0.01)
            f_eff = self._pipe_friction.get(seg, f_th)
            # If segment starts match previous end — add a duplicate with new values
            # to create a sharp step-change (shows DRA injection effect)
            if last_f_th is not None:
                pts.append({"km": km_s, "f_thermal": round(last_f_th, 7), "f_effective": round(last_f_eff, 7)})
            pts.append({"km": km_s,            "f_thermal": round(f_th, 7), "f_effective": round(f_eff, 7)})
            pts.append({"km": (km_s+km_e)//2,  "f_thermal": round(f_th, 7), "f_effective": round(f_eff, 7)})
            pts.append({"km": km_e,             "f_thermal": round(f_th, 7), "f_effective": round(f_eff, 7)})
            last_f_th, last_f_eff = f_th, f_eff
        return pts

    def _loop(self) -> None:
        info       = self.info
        node_names = info["junction_names"]
        pipe_names = info["pipe_names"]
        pump_names = info["pump_stations"]
        node_map   = info["node_map"]
        H_min, H_max = info["H_min_bar"], info["H_max_bar"]
        STEP_INTERVAL = 1.0

        while self.running:
            t0 = time.time()

            with self._lock:
                # ── MOC steps ────────────────────────────────────────────
                for _ in range(MOC_STEPS_PER_BROADCAST): self.solver.step()
                sim_t = self.solver.simulation_time

                # Advance valve schedules and push tau into MOC
                self._update_valves(sim_t)

                # ── Extract average flow for thermal model ───────────────
                flows_raw = self.solver.get_flows_m3s()
                avg_flow = np.mean([abs(v) for v in flows_raw.values()]) if flows_raw else 800.0
                avg_m3s  = max(avg_flow, 0.01)
                avg_kgs  = avg_m3s * info["density"]

                # ── Update thermal profile (every step) ──────────────────
                self._update_thermal_profile(mass_flow_kg_s=avg_kgs)

                # ── Read pressures ───────────────────────────────────────
                calc_p: Dict[str, float] = {}
                for name in node_names:
                    pi, ni = node_map[name]
                    H, _ = self.solver.get_state(pi, ni)
                    calc_p[name] = round(H * self.solver.pipes[pi].density * self.solver.pipes[pi].g / 1e5, 4)

                # ── Manual leak impact overlay for visible pressure/flow response ─────────
                self._apply_manual_leak_impact(sim_t, calc_p, flows_raw)

                # ── Flows ────────────────────────────────────────────────
                flows_m3s = {n: round(flows_raw.get(n,0.0),6) for n in pipe_names}
                flows_kgs = {n: round(v*info["density"],3) for n,v in flows_m3s.items()}

                # ── Sensor noise ─────────────────────────────────────────
                sensor_p   = {n: round(p+float(np.random.normal(0,PRESSURE_NOISE_STD)),4) for n,p in calc_p.items()}
                sensor_flow= {n: round(f+float(np.random.normal(0,FLOW_NOISE_STD)),6)     for n,f in flows_m3s.items()}

                pump_states   = self.get_pump_states()
                heater_states = self.get_heater_states()
                valve_states  = self.get_valve_states()
                dra_states    = self.get_dra_states()
                # Full profile for concentration chart (many points, includes spikes)
                dra_states["full_profile"] = self._dra.compute_full_profile()

                # ── Node payload ─────────────────────────────────────────
                nodes_payload = {
                    name: {
                        "pressure_bar"       : calc_p[name],
                        "sensor_pressure_bar": sensor_p[name],
                        "temperature_c"      : self._node_temps.get(name),
                        "status"             : ("alarm_low" if calc_p[name]<H_min else
                                                "alarm_high" if calc_p[name]>H_max else "normal"),
                        "is_pump"   : name in pump_names,
                        "is_relief" : name in info["relief_wells"],
                        "is_heater" : name in info["heating_stations"],
                        "has_leak"  : name in self._leaks,
                        "pump_on"   : pump_states.get(name,{}).get("on",False),
                        "heater_on" : heater_states.get(name,{}).get("on",False),
                    } for name in node_names
                }

                pipes_payload = {
                    name: {
                        "flow_m3s"           : flows_m3s[name],
                        "flow_kgs"           : flows_kgs[name],
                        "sensor_flow_m3s"    : sensor_flow[name],
                        "temperature_c"      : self._pipe_temps.get(name,29.0),
                        "friction_factor"    : round(self._pipe_friction.get(name,0.01),6),
                        "friction_factor_thermal": round(self._pipe_friction_thermal.get(name,0.01),6),
                    } for name in pipe_names
                }

                # ── Leak detection layer ─────────────────────────────────
                leak_detection = self._leak_detector.update(
                    sim_t=sim_t,
                    node_pressures_bar=sensor_p,
                    pipe_flows_kgs=flows_kgs,
                    pump_states=pump_states,
                    valve_states=valve_states,
                    active_leaks=self.get_leaks(),
                )

                payload = {
                    "pipeline_id"      : info["pipeline_id"],
                    "step"             : self.step_counter,
                    "simulated_time_s" : round(sim_t,2),
                    "nodes"            : nodes_payload,
                    "pipes"            : pipes_payload,
                    "pump_states"      : pump_states,
                    "heater_states"    : heater_states,
                    "junction_names"   : node_names,
                    "pipe_names"       : pipe_names,
                    "calc_pressure"    : [calc_p[n]    for n in node_names],
                    "sensor_pressure"  : [sensor_p[n]  for n in node_names],
                    "calc_flows"       : [flows_m3s[n] for n in pipe_names],
                    "calc_flows_kgs"   : [flows_kgs[n] for n in pipe_names],
                    "sensor_flows"     : [sensor_flow[n] for n in pipe_names],
                    # Temperature profile for distance-temperature chart
                    "temperature_profile": self._temp_profile,
                    "node_temperatures"  : self._node_temps,
                    "ambient_temp_c"     : self._ambient_temp_c,
                    "inlet_temp_c"       : self._inlet_temp_c,
                    "p_min_bar": H_min, "p_max_bar": H_max,
                    "active_leaks": self.get_leaks(),
                    "leak_nodes"  : list(self._leaks.keys()),
                    "leak_detection": leak_detection,
                    "dra_states"  : dra_states,
                    "valve_states"     : valve_states,
                    "paused"           : self._paused,
                    "safety_interlock" : self._safety_interlock,
                    "friction_profile": self._build_friction_profile(),
                }

                clean_payload = sanitize(payload)
                self.history = (self.history + [clean_payload])[-self._history_maxlen:]

                self._row_buffer.append(_make_row(
                    self.step_counter, sim_t, calc_p, sensor_p, flows_m3s,
                    self._pipe_temps, self._pipe_friction, pump_states,
                    node_names, pipe_names, pump_names,
                    heater_states, self._ambient_temp_c, self._inlet_temp_c, self._leaks
                ))
                if len(self._row_buffer) >= EXCEL_FLUSH_EVERY: self._flush_excel()

            if self.sio: asyncio.run(self._emit(clean_payload))
            self.step_counter += 1

            if self.step_counter % 25 == 0:
                avg_p = np.mean(list(calc_p.values()))
                pumps_on = [n for n,s in pump_states.items() if s["on"]]
                hs_on    = [k for k,v in heater_states.items() if v["on"]]
                print(f"  [Step {self.step_counter:>4} | T={sim_t:>7.1f}s]  "
                      f"avg P={avg_p:.2f} bar  pumps={pumps_on}  heaters={hs_on}")

            time.sleep(max(0.0, STEP_INTERVAL-(time.time()-t0)))

        print("\n🛑 Simulation loop exited.")

    async def _emit(self, payload):
        from socket_server import broadcast_live_update
        try:
            clean_payload = sanitize(payload)  # ✅ FIX: convert numpy → python types
            await broadcast_live_update(payload["pipeline_id"], clean_payload)
        except Exception as e:
            print(f"⚠️ Broadcast error: {e}")


# ── Singleton registry ────────────────────────────────────────────────────────

_services: Dict[str, SimulationService] = {}

def get_service(pid="alaska"):
    if pid not in _services: raise RuntimeError(f"Service '{pid}' not initialised.")
    return _services[pid]

def init_service(pid="alaska", sio=None):
    if pid in _services: return _services[pid]
    svc = SimulationService(sio=sio); _services[pid] = svc; return svc